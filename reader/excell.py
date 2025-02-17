from openpyxl import load_workbook

class Reader:
    def __init__(self, file):
        self.file = file
        pass
    
    def read(self):
        try:
            
            # Load the workbook
            workbook = load_workbook(filename=self.file)
            
            # Initialize results dictionary
            results = {}
            
            # Iterate through all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                # Iterate through all rows and columns
                for row in sheet.iter_rows():
                    for cell in row:
                        # Only include cells that have values
                        if cell.value is not None:
                            sheet_data.append({
                                'row': cell.row,
                                'column': cell.column_letter,
                                'value': str(cell.value)
                            })
                
                # Add sheet data to results if not empty
                if sheet_data:
                    results[sheet_name] = sheet_data
            
            return results
            
        except Exception as e:
            return {"error": f"Failed to read Excel file: {str(e)}"}


Reader.read("C:\Users\EfraimdeAlmeidaLima\OneDrive - Under Protection\Documentos\ATAS\ATA - CLIENTE - MOTIVO - D.A.TA.docx")